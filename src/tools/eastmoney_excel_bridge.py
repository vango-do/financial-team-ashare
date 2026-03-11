from __future__ import annotations

import json
import random
import re
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from functools import lru_cache
from pathlib import Path
from typing import Any

from src.tools.scrapling_framework import load_scrapling_fetchers

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


def _f(value: Any) -> float | None:
    if value in (None, "", "-"):
        return None
    try:
        out = float(value)
    except Exception:
        return None
    if out != out or out in (float("inf"), float("-inf")):
        return None
    return out


def _first(*vals: Any) -> Any:
    for val in vals:
        if val not in (None, "", "-"):
            return val
    return None


def _pct(num: Any, den: Any) -> float | None:
    n, d = _f(num), _f(den)
    if n is None or d in (None, 0):
        return None
    return n / d * 100.0


def _ratio(num: Any, den: Any) -> float | None:
    n, d = _f(num), _f(den)
    if n is None or d in (None, 0):
        return None
    return n / d


def _json_like(raw: str) -> dict[str, Any]:
    text = str(raw or "").strip()
    if not text:
        return {}
    if text[0] in "{[":
        return json.loads(text)
    m = re.search(r"^[^(]*\((.*)\)\s*;?\s*$", text, re.S)
    if m:
        return json.loads(m.group(1))
    return {}


def _resp_text(resp: Any) -> str:
    body = getattr(resp, "body", None)
    if isinstance(body, (bytes, bytearray)):
        for enc in ("utf-8", "gb18030", "gbk"):
            try:
                return bytes(body).decode(enc)
            except Exception:
                continue
        return bytes(body).decode("utf-8", errors="ignore")
    txt = getattr(resp, "text", None)
    return txt if isinstance(txt, str) else str(resp or "")


class EastmoneyDataCollector:
    _UA = (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )

    def __init__(self, stock_code: str):
        self.stock_code = str(stock_code or "").strip()
        if not re.fullmatch(r"\d{6}", self.stock_code):
            raise ValueError("股票代码必须是6位数字")
        tag, exchange, prefix = self._market_from_code(self.stock_code)
        self.market_tag = tag
        self.exchange = exchange
        self.secid = f"{prefix}.{self.stock_code}"
        self.secucode = f"{self.stock_code}.{self.market_tag}"
        self.max_retries = 5
        self.request_timeout = 20
        self._cache: dict[str, Any] = {}
        self._lock = threading.Lock()
        self._fetcher = load_scrapling_fetchers().get("requests")

    @staticmethod
    def _market_from_code(code: str) -> tuple[str, str, str]:
        if code.startswith(("6", "9", "5")):
            return "SH", "SSE", "1"
        if code.startswith(("4", "8")):
            return "BJ", "BSE", "0"
        return "SZ", "SZSE", "0"

    @staticmethod
    def _board(code: str) -> str:
        if code.startswith("688"):
            return "科创板"
        if code.startswith("300"):
            return "创业板"
        if code.startswith(("4", "8")):
            return "北交所"
        return "主板"

    @staticmethod
    def _date(value: Any) -> str | None:
        if value is None:
            return None
        txt = str(value).strip()
        return txt[:10] if txt else None

    def _cache_key(self, url: str, params: dict[str, Any] | None, mode: str) -> str:
        return f"{mode}|{url}|{json.dumps(params or {}, sort_keys=True, ensure_ascii=False)}"

    def _request(self, url: str, params: dict[str, Any] | None = None, mode: str = "json") -> Any:
        key = self._cache_key(url, params, mode)
        with self._lock:
            if key in self._cache:
                return self._cache[key]

        headers = {
            "User-Agent": self._UA,
            "Referer": f"https://data.eastmoney.com/bbsj/{self.stock_code}.html",
            "Accept": "*/*",
        }
        last_error: Exception | None = None
        for i in range(self.max_retries):
            if i > 0:
                time.sleep(min(0.35 * (2 ** (i - 1)) + random.uniform(0.05, 0.2), 3.0))
            try:
                if self._fetcher is None:
                    import requests

                    resp = requests.get(url, params=params or {}, headers=headers, timeout=self.request_timeout)
                    raw = resp.text
                    obj = raw if mode == "text" else (_json_like(raw) if resp.headers.get("content-type", "").find("json") < 0 else resp.json())
                else:
                    resp = self._fetcher.get(
                        url,
                        params=params or {},
                        headers=headers,
                        timeout=self.request_timeout,
                        retries=1,
                        impersonate=random.choice(("chrome", "edge")),
                        stealthy_headers=True,
                    )
                    raw = _resp_text(resp)
                    obj = raw if mode == "text" else (_json_like(raw) if not raw.strip().startswith("{") else json.loads(raw))
                with self._lock:
                    self._cache[key] = obj
                return obj
            except Exception as exc:  # pragma: no cover
                last_error = exc
        if last_error is not None:
            raise last_error
        return {} if mode == "json" else ""

    def _json(self, url: str, params: dict[str, Any] | None = None) -> dict[str, Any]:
        out = self._request(url, params=params, mode="json")
        return out if isinstance(out, dict) else {}

    def _text(self, url: str, params: dict[str, Any] | None = None) -> str:
        return str(self._request(url, params=params, mode="text") or "")

    def _quote(self) -> dict[str, Any]:
        fields = ",".join(["f57", "f58", "f59", "f43", "f44", "f45", "f46", "f47", "f48", "f60", "f116", "f117", "f162", "f167", "f173", "f184", "f185", "f186", "f187", "f84", "f85"])
        data = self._json("https://push2.eastmoney.com/api/qt/stock/get", {"fltt": "2", "invt": "2", "fields": fields, "secid": self.secid})
        return (data.get("data") or {}) if isinstance(data, dict) else {}

    def _finance(self, type_name: str) -> dict[str, Any]:
        data = self._json("https://datacenter.eastmoney.com/securities/api/data/get", {"type": type_name, "sty": "ALL", "filter": f'(SECUCODE="{self.secucode}")', "p": 1, "ps": 1, "sr": -1, "st": "REPORT_DATE", "source": "HSF10", "client": "PC"})
        rows = ((data.get("result") or {}).get("data") or []) if isinstance(data, dict) else []
        return rows[0] if rows else {}

    def _notice(self) -> dict[str, Any]:
        data = self._json("https://np-anotice-stock.eastmoney.com/api/security/ann", {"ann_type": "A", "client_source": "web", "stock_list": self.stock_code, "page_size": 10, "page_index": 1})
        rows = ((data.get("data") or {}).get("list") or []) if isinstance(data, dict) else []
        return rows[0] if rows else {}

    def _insider(self) -> dict[str, Any]:
        data = self._json("https://datacenter-web.eastmoney.com/api/data/v1/get", {"reportName": "RPT_SHARE_HOLDER_INCREASE", "columns": "ALL", "filter": f'(SECURITY_CODE="{self.stock_code}")', "pageNumber": 1, "pageSize": 1, "sortTypes": -1, "sortColumns": "END_DATE", "source": "WEB", "client": "WEB"})
        rows = ((data.get("result") or {}).get("data") or []) if isinstance(data, dict) else []
        return rows[0] if rows else {}

    def build_field_values(self) -> dict[str, Any]:
        tasks = {
            "quote": self._quote,
            "income": lambda: self._finance("RPT_F10_FINANCE_GINCOME"),
            "cash": lambda: self._finance("RPT_F10_FINANCE_GCASHFLOW"),
            "balance": lambda: self._finance("RPT_F10_FINANCE_GBALANCE"),
            "notice": self._notice,
            "insider": self._insider,
        }
        res: dict[str, Any] = {}
        with ThreadPoolExecutor(max_workers=4) as pool:
            m = {pool.submit(fn): k for k, fn in tasks.items()}
            for future in as_completed(m):
                key = m[future]
                try:
                    res[key] = future.result()
                except Exception:
                    res[key] = {}
        q, p, c, b, n, i = res["quote"], res["income"], res["cash"], res["balance"], res["notice"], res["insider"]
        revenue = _first(p.get("TOTAL_OPERATE_INCOME"), q.get("f183"))
        net_income = _first(p.get("PARENT_NETPROFIT"), p.get("NETPROFIT"))
        op_income = _first(p.get("OPERATE_PROFIT"), p.get("TOTAL_PROFIT"))
        op_cost = p.get("OPERATE_COST")
        gross_profit = (_f(revenue) - _f(op_cost)) if _f(revenue) is not None and _f(op_cost) is not None else None
        eq = _first(b.get("TOTAL_PARENT_EQUITY"), b.get("TOTAL_EQUITY"))
        assets = b.get("TOTAL_ASSETS")
        liabilities = b.get("TOTAL_LIABILITIES")
        current_assets = b.get("TOTAL_CURRENT_ASSETS")
        current_liabilities = b.get("TOTAL_CURRENT_LIAB")
        shares = _first(b.get("SHARE_CAPITAL"), q.get("f84"), q.get("f85"))
        capex = _f(c.get("CONSTRUCT_LONG_ASSET"))
        op_cash = _f(c.get("NETCASH_OPERATE"))
        fcf = (op_cash - capex) if op_cash is not None and capex is not None else None
        debt = (_f(b.get("SHORT_LOAN")) or 0.0) + (_f(b.get("LONG_LOAN")) or 0.0) + (_f(b.get("BOND_PAYABLE")) or 0.0)
        eps = p.get("BASIC_EPS")
        pe = q.get("f162")
        eg = _f(_first(p.get("PARENT_NETPROFIT_YOY"), q.get("f185")))
        peg = (_f(pe) / eg) if _f(pe) is not None and eg not in (None, 0) and eg > 0 else None
        values = {
            "stock_code": self.stock_code,
            "stock_name": str(_first(q.get("f58"), self.stock_code)),
            "exchange": self.exchange,
            "board": self._board(self.stock_code),
            "industry": "未知行业",
            "report_period": self._date(_first(p.get("REPORT_DATE"), b.get("REPORT_DATE"))),
            "price_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "current_price": q.get("f43"),
            "price_open": q.get("f46"),
            "price_high": q.get("f44"),
            "price_low": q.get("f45"),
            "price_close": _first(q.get("f43"), q.get("f60")),
            "volume": q.get("f47"),
            "amount": q.get("f48"),
            "market_cap_total": q.get("f116"),
            "pe_ttm": pe,
            "pb": q.get("f167"),
            "return_on_equity": _pct(net_income, eq),
            "return_on_invested_capital": _pct(op_income, assets),
            "debt_to_equity": _ratio(liabilities, eq),
            "operating_margin": _pct(op_income, revenue),
            "gross_margin": _first(_pct(gross_profit, revenue), q.get("f186")),
            "net_margin": _first(_pct(net_income, revenue), q.get("f187")),
            "current_ratio": _ratio(current_assets, current_liabilities),
            "asset_turnover": _ratio(revenue, assets),
            "revenue": revenue,
            "net_income": net_income,
            "free_cash_flow": fcf,
            "capital_expenditure": capex,
            "depreciation_and_amortization": (_f(c.get("FA_IR_DEPR")) or 0.0) + (_f(c.get("IR_DEPR")) or 0.0) + (_f(c.get("IA_AMORTIZE")) or 0.0),
            "total_assets": assets,
            "total_liabilities": liabilities,
            "shareholders_equity": eq,
            "current_assets": current_assets,
            "current_liabilities": current_liabilities,
            "outstanding_shares": shares,
            "dividends_and_other_cash_distributions": c.get("ASSIGN_DIVIDEND_PORFIT"),
            "issuance_or_purchase_of_equity_shares": _first(c.get("ACCEPT_INVEST_CASH"), c.get("BUY_SUBSIDIARY_EQUITY")),
            "gross_profit": gross_profit,
            "intrinsic_value_estimate": (_ratio(eq, shares) * 1.5) if _ratio(eq, shares) is not None else None,
            "earnings_per_share": eps,
            "operating_income": op_income,
            "cash_and_equivalents": b.get("MONETARYFUNDS"),
            "total_debt": debt,
            "ebit": _f(op_income),
            "ebitda": (_f(op_income) + ((_f(c.get("FA_IR_DEPR")) or 0.0) + (_f(c.get("IR_DEPR")) or 0.0) + (_f(c.get("IA_AMORTIZE")) or 0.0))) if _f(op_income) is not None else None,
            "company_news_title": n.get("title"),
            "company_news_publish_time": self._date(_first(n.get("display_time"), n.get("notice_date"))),
            "company_news_source": ((n.get("columns") or [{}])[0] or {}).get("column_name") if isinstance(n.get("columns"), list) else n.get("source_type"),
            "insider_transaction_type": i.get("DIRECTION"),
            "insider_transaction_shares": _first(i.get("CHANGE_NUM_SYMBOL"), i.get("CHANGE_NUM")),
            "insider_transaction_time": self._date(_first(i.get("END_DATE"), i.get("TRADE_DATE"), i.get("NOTICE_DATE"))),
            "revenue_growth": _first(p.get("TOTAL_OPERATE_INCOME_YOY"), q.get("f184")),
            "earnings_growth": _first(p.get("PARENT_NETPROFIT_YOY"), q.get("f185")),
            "peg_ratio": peg,
            "research_and_development": p.get("RESEARCH_EXPENSE"),
            "goodwill_and_intangible_assets": (_f(b.get("GOODWILL")) or 0.0) + (_f(b.get("INTANGIBLE_ASSET")) or 0.0),
        }
        return values


@lru_cache(maxsize=256)
def collect_field_values(stock_code: str) -> dict[str, Any]:
    return EastmoneyDataCollector(stock_code).build_field_values()


_AGENT_TO_SHEET = {
    "warren_buffett_agent": "沃伦·巴菲特",
    "stanley_druckenmiller_agent": "斯坦利·德鲁肯米勒",
    "fundamentals_analyst_agent": "基本面分析师",
    "growth_analyst_agent": "成长风格分析师",
    "peter_lynch_agent": "彼得·林奇",
    "charlie_munger_agent": "查理·芒格",
    "soros_agent": "乔治·索罗斯",
}


@lru_cache(maxsize=1)
def _sheet_fields() -> dict[str, set[str]]:
    if load_workbook is None:
        return {}
    root = Path(__file__).resolve().parents[2]
    for p in [root / "数据获取" / "专家分析数据模板.xlsx", root / "专家分析数据模板.xlsx"]:
        if not p.exists():
            continue
        try:
            wb = load_workbook(p, read_only=True)
            out: dict[str, set[str]] = {}
            for ws in wb.worksheets:
                out[ws.title] = {
                    str(row[2]).strip()
                    for row in ws.iter_rows(min_row=2, values_only=True)
                    if len(row) > 2 and row[2]
                }
            return out
        except Exception:
            continue
    return {}


def get_allowed_fields_for_agent(agent_name: str | None) -> set[str] | None:
    if not agent_name:
        return None
    sheet = _AGENT_TO_SHEET.get(agent_name)
    if not sheet:
        return None
    mapping = _sheet_fields()
    return set(mapping.get(sheet) or []) or None


def filter_values_for_agent(values: dict[str, Any], agent_name: str | None) -> dict[str, Any]:
    allow = get_allowed_fields_for_agent(agent_name)
    if not allow:
        return dict(values)
    return {k: values.get(k) for k in allow}

