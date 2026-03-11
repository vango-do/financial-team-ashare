from __future__ import annotations

import json
import math
import os
import random
import re
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import quote

import openpyxl

def _candidate_scrapling_paths() -> list[Path]:
    project_root = Path(__file__).resolve().parents[1]
    env_path = os.getenv("SCRAPLING_CN_LIB") or os.getenv("SCRAPLING_HOME")
    paths: list[Path] = []
    if env_path:
        paths.append(Path(env_path).expanduser())
    paths.extend(
        [
            project_root / "scrapling_cn_lib",
            project_root / "vendor" / "scrapling",
            project_root / "third_party" / "scrapling",
        ]
    )
    return paths


for _candidate in _candidate_scrapling_paths():
    if (_candidate / "scrapling" / "__init__.py").exists():
        if str(_candidate) not in sys.path:
            sys.path.insert(0, str(_candidate))
        break

from scrapling.fetchers import Fetcher  # type: ignore  # noqa: E402


def parse_json_like(raw: str) -> dict[str, Any]:
    raw = (raw or "").strip()
    if not raw:
        return {}
    if raw[0] in "{[":
        return json.loads(raw)
    m = re.search(r"^[^(]*\((.*)\)\s*;?\s*$", raw, re.S)
    if m:
        return json.loads(m.group(1))
    raise ValueError("Unable to parse JSON/JSONP response.")


def safe_float(value: Any) -> float | None:
    if value is None or value == "" or value == "-":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def first_not_none(*values: Any) -> Any:
    for v in values:
        if v is not None and v != "" and v != "-":
            return v
    return None


def pct(num: Any, den: Any) -> float | None:
    n = safe_float(num)
    d = safe_float(den)
    if n is None or d is None or d == 0:
        return None
    return n / d * 100


def ratio(num: Any, den: Any) -> float | None:
    n = safe_float(num)
    d = safe_float(den)
    if n is None or d is None or d == 0:
        return None
    return n / d


def round_or_none(value: Any, digits: int = 4) -> float | None:
    v = safe_float(value)
    if v is None:
        return None
    return round(v, digits)


def sanitize_filename(name: str) -> str:
    name = (name or "").strip()
    name = re.sub(r"[\\/:*?\"<>|]", "_", name)
    name = re.sub(r"\s+", "", name)
    return name or "输出结果"


class EastmoneyDataCollector:
    def __init__(self, stock_code: str):
        self.stock_code = stock_code.strip()
        if not re.fullmatch(r"\d{6}", self.stock_code):
            raise ValueError("股票代码必须是6位数字。")

        self.market_tag, self.exchange, self.secid_prefix = self._market_from_code(self.stock_code)
        self.secid = f"{self.secid_prefix}.{self.stock_code}"
        self.secucode = f"{self.stock_code}.{self.market_tag}"
        self.request_timeout = 20
        self.max_retries = 5
        self.base_backoff = 0.35
        self.max_backoff = 3.0
        # Keep concurrency moderate to speed up while lowering anti-crawl risk.
        self.max_workers = 4
        self.impersonate_pool = ("chrome", "edge")
        self._cache: dict[str, Any] = {}
        self._cache_lock = threading.Lock()
        self.eastmoney_token = os.getenv("EASTMONEY_TOKEN", "").strip()

    @staticmethod
    def _market_from_code(code: str) -> tuple[str, str, str]:
        if code.startswith(("6", "9", "5")):
            return "SH", "SSE", "1"
        if code.startswith(("0", "2", "3", "1")):
            return "SZ", "SZSE", "0"
        if code.startswith(("4", "8")):
            return "BJ", "BSE", "0"
        return "SZ", "SZSE", "0"

    @staticmethod
    def board_from_code(code: str) -> str:
        if code.startswith("688"):
            return "科创板"
        if code.startswith("300"):
            return "创业板"
        if code.startswith(("4", "8")):
            return "北交所"
        return "主板"

    def _build_cache_key(self, url: str, params: dict[str, Any] | None, parse_mode: str) -> str:
        payload = json.dumps(params or {}, sort_keys=True, ensure_ascii=False, default=str)
        return f"{parse_mode}|{url}|{payload}"

    @staticmethod
    def _looks_like_block(raw_text: str, status: int) -> bool:
        if status in (403, 429, 503):
            return True
        sample = (raw_text or "")[:4000]
        block_markers = (
            "访问过于频繁",
            "访问频繁",
            "forbidden",
            "verify",
            "captcha",
            "waf",
        )
        lower = sample.lower()
        return any(m in sample or m in lower for m in block_markers)

    def _request(self, url: str, params: dict[str, Any] | None = None, parse_mode: str = "json") -> Any:
        cache_key = self._build_cache_key(url, params, parse_mode)
        with self._cache_lock:
            if cache_key in self._cache:
                return self._cache[cache_key]

        last_err: Exception | None = None
        request_params = params or {}
        headers = {
            "Accept": "*/*",
            "Accept-Language": "zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7",
            "Cache-Control": "no-cache",
            "Pragma": "no-cache",
        }

        for attempt in range(self.max_retries):
            if attempt > 0:
                backoff = min(self.base_backoff * (2 ** (attempt - 1)) + random.uniform(0.05, 0.25), self.max_backoff)
                time.sleep(backoff)
            # Small jitter lowers burst pattern and reduces anti-crawl hit rate.
            time.sleep(random.uniform(0.02, 0.08))

            try:
                resp = Fetcher.get(
                    url,
                    params=request_params,
                    headers=headers,
                    timeout=self.request_timeout,
                    retries=1,
                    impersonate=random.choice(self.impersonate_pool),
                    stealthy_headers=True,
                )
                raw = resp.body.decode("utf-8", "ignore")
                if self._looks_like_block(raw, getattr(resp, "status", 0)):
                    raise RuntimeError(f"blocked_or_rate_limited: {getattr(resp, 'status', 'unknown')}")

                if parse_mode == "text":
                    result: Any = raw
                else:
                    try:
                        result = resp.json()
                    except Exception:
                        result = parse_json_like(raw)

                with self._cache_lock:
                    self._cache[cache_key] = result
                return result
            except Exception as exc:  # pragma: no cover - network dependent
                last_err = exc

        if last_err is not None:
            raise last_err
        return {}

    def request_json(self, url: str, params: dict[str, Any] | None = None) -> dict[str, Any]:
        data = self._request(url, params=params, parse_mode="json")
        return data if isinstance(data, dict) else {}

    def get_quote(self) -> dict[str, Any]:
        fields = ",".join(
            [
                "f57",
                "f58",
                "f59",
                "f43",
                "f44",
                "f45",
                "f46",
                "f47",
                "f48",
                "f60",
                "f116",
                "f117",
                "f162",
                "f167",
                "f168",
                "f169",
                "f170",
                "f173",
                "f84",
                "f85",
                "f183",
                "f184",
                "f185",
                "f186",
                "f187",
                "f188",
                "f190",
            ]
        )
        data = self.request_json(
            "https://push2.eastmoney.com/api/qt/stock/get",
            params={"fltt": "2", "invt": "2", "fields": fields, "secid": self.secid},
        )
        return (data.get("data") or {}) if isinstance(data, dict) else {}

    def get_finance_row(self, type_name: str) -> dict[str, Any]:
        params = {
            "type": type_name,
            "sty": "ALL",
            "filter": f'(SECUCODE="{self.secucode}")',
            "p": 1,
            "ps": 1,
            "sr": -1,
            "st": "REPORT_DATE",
            "source": "HSF10",
            "client": "PC",
        }
        data = self.request_json("https://datacenter.eastmoney.com/securities/api/data/get", params=params)
        rows = ((data.get("result") or {}).get("data") or []) if isinstance(data, dict) else []
        return rows[0] if rows else {}

    def get_flow_snapshot(self) -> dict[str, Any]:
        fields = ",".join(
            [
                "f62",
                "f184",
                "f66",
                "f69",
                "f72",
                "f75",
                "f78",
                "f81",
                "f84",
                "f87",
                "f64",
                "f65",
                "f70",
                "f71",
                "f76",
                "f77",
                "f82",
                "f83",
                "f164",
                "f166",
                "f168",
                "f170",
                "f172",
                "f252",
                "f253",
                "f254",
                "f255",
                "f256",
                "f124",
                "f6",
                "f278",
                "f279",
                "f280",
                "f281",
                "f282",
            ]
        )
        data = self.request_json(
            "https://push2.eastmoney.com/api/qt/ulist.np/get",
            params={
                "fltt": "2",
                "secids": self.secid,
                "fields": fields,
                "ut": "b2884a393a59ad64002292a3e90d46a5",
            },
        )
        rows = ((data.get("data") or {}).get("diff") or []) if isinstance(data, dict) else []
        return rows[0] if rows else {}

    def get_stock_info(self) -> dict[str, Any]:
        page = self.request_text(f"https://data.eastmoney.com/notices/stock/{self.stock_code}.html")
        m = re.search(r"stockInfo\s*=\s*(\{.*?\});", page, re.S)
        if not m:
            return {}
        try:
            return json.loads(m.group(1))
        except Exception:
            return {}

    def request_text(self, url: str, params: dict[str, Any] | None = None) -> str:
        text = self._request(url, params=params, parse_mode="text")
        return str(text or "")

    def get_latest_notice(self) -> dict[str, Any]:
        data = self.request_json(
            "https://np-anotice-stock.eastmoney.com/api/security/ann",
            params={
                "ann_type": "A",
                "client_source": "web",
                "stock_list": self.stock_code,
                "page_size": 10,
                "page_index": 1,
            },
        )
        rows = ((data.get("data") or {}).get("list") or []) if isinstance(data, dict) else []
        return rows[0] if rows else {}

    def get_latest_insider_trade(self) -> dict[str, Any]:
        data = self.request_json(
            "https://datacenter-web.eastmoney.com/api/data/v1/get",
            params={
                "reportName": "RPT_SHARE_HOLDER_INCREASE",
                "columns": "ALL",
                "filter": f'(SECURITY_CODE="{self.stock_code}")',
                "pageNumber": 1,
                "pageSize": 1,
                "sortTypes": -1,
                "sortColumns": "END_DATE",
                "source": "WEB",
                "client": "WEB",
            },
        )
        rows = ((data.get("result") or {}).get("data") or []) if isinstance(data, dict) else []
        return rows[0] if rows else {}

    def get_northbound_inflow_20d(self) -> float | None:
        if not self.eastmoney_token:
            raise RuntimeError(
                "缺少 EASTMONEY_TOKEN。请在环境变量中配置 EASTMONEY_TOKEN 后重试。"
            )
        data = self.request_json(
            "https://datacenter-web.eastmoney.com/securities/api/data/v1/get",
            params={
                "reportName": "RPT_MUTUAL_NETINFLOW_DETAILS",
                "columns": "DIRECTION_TYPE,TRADE_DATE,NET_INFLOW_BOTH,TIME_TYPE",
                "token": self.eastmoney_token,
                "client": "WEB",
                "filter": '(DIRECTION_TYPE="1")(TIME_TYPE="4")',
                "sortColumns": "TRADE_DATE",
                "sortTypes": "-1",
                "pageNumber": 1,
                "pageSize": 20,
            },
        )
        rows = ((data.get("result") or {}).get("data") or []) if isinstance(data, dict) else []
        if not rows:
            return None
        vals = [safe_float(r.get("NET_INFLOW_BOTH")) or 0.0 for r in rows[:20]]
        return round(sum(vals), 2)

    def get_northbound_hold_change_20d(self) -> tuple[float | None, dict[str, Any] | None]:
        data = self.request_json(
            "https://datacenter-web.eastmoney.com/api/data/v1/get",
            params={
                "reportName": "RPT_MUTUAL_HOLDSTOCKNDATE_STA",
                "columns": "SECUCODE,SECURITY_CODE,SECURITY_NAME,TRADE_DATE,HOLD_MARKET_CAP,HOLD_SHARES,CLOSE_PRICE,CHANGE_RATE",
                "filter": f'(INTERVAL_TYPE="1")(SECUCODE="{self.secucode}")',
                "pageNumber": 1,
                "pageSize": 60,
                "sortTypes": -1,
                "sortColumns": "TRADE_DATE",
                "source": "WEB",
                "client": "WEB",
            },
        )
        rows = ((data.get("result") or {}).get("data") or []) if isinstance(data, dict) else []
        if not rows:
            return None, None
        latest = rows[0]
        idx = min(19, len(rows) - 1)
        latest_shares = safe_float(latest.get("HOLD_SHARES"))
        old_shares = safe_float(rows[idx].get("HOLD_SHARES"))
        if latest_shares is None or old_shares is None:
            return None, latest
        return round(latest_shares - old_shares, 2), latest

    @staticmethod
    def parse_date_text(value: Any) -> str | None:
        if value is None:
            return None
        txt = str(value).strip()
        if not txt:
            return None
        return txt[:10] if len(txt) >= 10 else txt

    def build_field_values(self) -> dict[str, Any]:
        if not self.eastmoney_token:
            raise RuntimeError(
                "缺少 EASTMONEY_TOKEN。请在环境变量中配置 EASTMONEY_TOKEN 后重试。"
            )

        tasks: dict[str, Any] = {
            "quote": self.get_quote,
            "income": lambda: self.get_finance_row("RPT_F10_FINANCE_GINCOME"),
            "cash": lambda: self.get_finance_row("RPT_F10_FINANCE_GCASHFLOW"),
            "balance": lambda: self.get_finance_row("RPT_F10_FINANCE_GBALANCE"),
            "flow": self.get_flow_snapshot,
            "stock_info": self.get_stock_info,
            "notice": self.get_latest_notice,
            "insider": self.get_latest_insider_trade,
            "north_inflow_20d": self.get_northbound_inflow_20d,
            "north_hold_change_20d": self.get_northbound_hold_change_20d,
        }
        defaults: dict[str, Any] = {
            "quote": {},
            "income": {},
            "cash": {},
            "balance": {},
            "flow": {},
            "stock_info": {},
            "notice": {},
            "insider": {},
            "north_inflow_20d": None,
            "north_hold_change_20d": (None, None),
        }
        results: dict[str, Any] = {}
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            future_map = {executor.submit(fn): key for key, fn in tasks.items()}
            for future in as_completed(future_map):
                key = future_map[future]
                try:
                    results[key] = future.result()
                except Exception:
                    results[key] = defaults[key]

        quote = results["quote"]
        income = results["income"]
        cash = results["cash"]
        balance = results["balance"]
        flow = results["flow"]
        stock_info = results["stock_info"]
        notice = results["notice"]
        insider = results["insider"]
        north_inflow_20d = results["north_inflow_20d"]
        north_hold_change_20d, north_latest_detail = results["north_hold_change_20d"]

        stock_name = str(first_not_none(quote.get("f58"), stock_info.get("name"), self.stock_code))
        industry = first_not_none(stock_info.get("hyname"), "未知行业")
        board = self.board_from_code(self.stock_code)

        revenue = first_not_none(income.get("TOTAL_OPERATE_INCOME"), quote.get("f183"))
        operating_profit = first_not_none(income.get("OPERATE_PROFIT"), income.get("TOTAL_PROFIT"))
        operate_cost = income.get("OPERATE_COST")
        net_income = first_not_none(income.get("PARENT_NETPROFIT"), income.get("NETPROFIT"))
        gross_profit = None
        if safe_float(revenue) is not None and safe_float(operate_cost) is not None:
            gross_profit = safe_float(revenue) - safe_float(operate_cost)  # type: ignore[operator]

        total_assets = balance.get("TOTAL_ASSETS")
        total_liabilities = balance.get("TOTAL_LIABILITIES")
        shareholders_equity = first_not_none(balance.get("TOTAL_PARENT_EQUITY"), balance.get("TOTAL_EQUITY"))
        current_assets = balance.get("TOTAL_CURRENT_ASSETS")
        current_liabilities = balance.get("TOTAL_CURRENT_LIAB")
        inventory = balance.get("INVENTORY")
        accounts_receivable = first_not_none(balance.get("ACCOUNTS_RECE"), balance.get("NOTE_ACCOUNTS_RECE"))

        short_loan = safe_float(balance.get("SHORT_LOAN")) or 0.0
        long_loan = safe_float(balance.get("LONG_LOAN")) or 0.0
        bonds_payable = safe_float(balance.get("BOND_PAYABLE")) or 0.0
        total_debt = short_loan + long_loan + bonds_payable

        capex = safe_float(cash.get("CONSTRUCT_LONG_ASSET"))
        op_cash = safe_float(cash.get("NETCASH_OPERATE"))
        depr_amort = (
            (safe_float(cash.get("FA_IR_DEPR")) or 0.0)
            + (safe_float(cash.get("IR_DEPR")) or 0.0)
            + (safe_float(cash.get("IA_AMORTIZE")) or 0.0)
            + (safe_float(cash.get("OILGAS_BIOLOGY_DEPR")) or 0.0)
            + (safe_float(cash.get("USERIGHT_ASSET_AMORTIZE")) or 0.0)
        )
        free_cash_flow = (op_cash - capex) if op_cash is not None and capex is not None else None

        ebit = safe_float(operating_profit)
        ebitda = (ebit + depr_amort) if ebit is not None else None

        shares = first_not_none(balance.get("SHARE_CAPITAL"), quote.get("f84"), quote.get("f85"))
        net_assets_per_share = ratio(shareholders_equity, shares)
        intrinsic_value = (net_assets_per_share * 1.5) if net_assets_per_share is not None else None

        revenue_growth = first_not_none(income.get("TOTAL_OPERATE_INCOME_YOY"), quote.get("f184"))
        earnings_growth = first_not_none(income.get("PARENT_NETPROFIT_YOY"), quote.get("f185"))
        pe_ttm = quote.get("f162")
        peg_ratio = None
        if safe_float(pe_ttm) is not None and safe_float(earnings_growth) not in (None, 0):
            eg = safe_float(earnings_growth)
            if eg is not None and eg > 0:
                peg_ratio = safe_float(pe_ttm) / eg

        gross_margin = first_not_none(pct(gross_profit, revenue), quote.get("f186"))
        net_margin = first_not_none(pct(net_income, revenue), quote.get("f187"))
        operating_margin = pct(operating_profit, revenue)
        current_ratio = ratio(current_assets, current_liabilities)
        debt_to_equity = ratio(total_liabilities, shareholders_equity)
        asset_turnover = ratio(revenue, total_assets)
        inventory_turnover = ratio(operate_cost, inventory)
        receivables_turnover = ratio(revenue, accounts_receivable)
        roe = first_not_none(pct(net_income, shareholders_equity), quote.get("f173"))
        roic = pct(ebit, total_assets)

        main_flow = safe_float(flow.get("f62"))
        market_cap = safe_float(quote.get("f116"))
        theme_hot_money_index = None
        if main_flow is not None and market_cap and market_cap != 0:
            theme_hot_money_index = main_flow / market_cap * 10000

        news_title = notice.get("title") or ""
        positive_kw = ("增持", "回购", "中标", "签约", "预增", "重大利好")
        negative_kw = ("减持", "处罚", "诉讼", "风险", "亏损", "违约", "退市")
        if any(k in str(news_title) for k in positive_kw):
            policy_signal = "偏多"
        elif any(k in str(news_title) for k in negative_kw):
            policy_signal = "偏空"
        else:
            policy_signal = "中性"

        if north_inflow_20d is None:
            national_team_signal = "暂无显著信号"
        elif north_inflow_20d > 0:
            national_team_signal = "偏多（北向资金净流入）"
        elif north_inflow_20d < 0:
            national_team_signal = "偏空（北向资金净流出）"
        else:
            national_team_signal = "中性"

        news_source = None
        columns = notice.get("columns") if isinstance(notice, dict) else None
        if isinstance(columns, list) and columns:
            first_col = columns[0] or {}
            news_source = first_col.get("column_name")
        if news_source is None:
            news_source = first_not_none(notice.get("source_type"), "东方财富")

        values: dict[str, Any] = {
            "stock_code": self.stock_code,
            "stock_name": stock_name,
            "exchange": self.exchange,
            "board": board,
            "industry": industry,
            "report_period": self.parse_date_text(first_not_none(income.get("REPORT_DATE"), balance.get("REPORT_DATE"))),
            "price_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "current_price": quote.get("f43"),
            "price_open": quote.get("f46"),
            "price_high": quote.get("f44"),
            "price_low": quote.get("f45"),
            "price_close": first_not_none(quote.get("f43"), quote.get("f60")),
            "volume": quote.get("f47"),
            "amount": quote.get("f48"),
            "market_cap_total": quote.get("f116"),
            "pe_ttm": pe_ttm,
            "pb": quote.get("f167"),
            "return_on_equity": round_or_none(roe, 4),
            "return_on_invested_capital": round_or_none(roic, 4),
            "debt_to_equity": round_or_none(debt_to_equity, 6),
            "operating_margin": round_or_none(operating_margin, 4),
            "gross_margin": round_or_none(gross_margin, 4),
            "current_ratio": round_or_none(current_ratio, 6),
            "asset_turnover": round_or_none(asset_turnover, 6),
            "revenue": revenue,
            "net_income": net_income,
            "free_cash_flow": round_or_none(free_cash_flow, 2),
            "capital_expenditure": capex,
            "depreciation_and_amortization": round_or_none(depr_amort, 2),
            "total_assets": total_assets,
            "total_liabilities": total_liabilities,
            "shareholders_equity": shareholders_equity,
            "current_assets": current_assets,
            "current_liabilities": current_liabilities,
            "outstanding_shares": shares,
            "dividends_and_other_cash_distributions": cash.get("ASSIGN_DIVIDEND_PORFIT"),
            "issuance_or_purchase_of_equity_shares": first_not_none(cash.get("ACCEPT_INVEST_CASH"), cash.get("BUY_SUBSIDIARY_EQUITY")),
            "gross_profit": round_or_none(gross_profit, 2),
            "intrinsic_value_estimate": round_or_none(intrinsic_value, 4),
            "earnings_per_share": income.get("BASIC_EPS"),
            "operating_income": operating_profit,
            "cash_and_equivalents": balance.get("MONETARYFUNDS"),
            "total_debt": round_or_none(total_debt, 2),
            "ebit": round_or_none(ebit, 2),
            "ebitda": round_or_none(ebitda, 2),
            "company_news_title": news_title,
            "company_news_publish_time": self.parse_date_text(first_not_none(notice.get("display_time"), notice.get("notice_date"))),
            "company_news_source": news_source,
            "insider_transaction_type": insider.get("DIRECTION"),
            "insider_transaction_shares": first_not_none(insider.get("CHANGE_NUM_SYMBOL"), insider.get("CHANGE_NUM")),
            "insider_transaction_time": self.parse_date_text(first_not_none(insider.get("END_DATE"), insider.get("TRADE_DATE"), insider.get("NOTICE_DATE"))),
            "net_margin": round_or_none(net_margin, 4),
            "revenue_growth": round_or_none(revenue_growth, 4),
            "earnings_growth": round_or_none(earnings_growth, 4),
            "northbound_net_inflow_20d": north_inflow_20d,
            "northbound_holding_change_20d": north_hold_change_20d,
            "national_team_signal": national_team_signal,
            "policy_signal": policy_signal,
            "theme_hot_money_index": round_or_none(theme_hot_money_index, 6),
            "peg_ratio": round_or_none(peg_ratio, 6),
            "research_and_development": income.get("RESEARCH_EXPENSE"),
            "goodwill_and_intangible_assets": round_or_none(
                (safe_float(balance.get("GOODWILL")) or 0.0) + (safe_float(balance.get("INTANGIBLE_ASSET")) or 0.0),
                2,
            ),
        }

        # Extra placeholders for fields that are present in some versions/sheets.
        if north_latest_detail:
            values["northbound_holding_detail"] = (
                f"最新持股{north_latest_detail.get('HOLD_SHARES')}股, 日期{self.parse_date_text(north_latest_detail.get('TRADE_DATE'))}"
            )
        values["northbound_top_trades"] = "见沪深港通十大成交股榜单"

        return values


def find_template_excel(cwd: Path) -> Path:
    preferred = cwd / "专家分析数据模板.xlsx"
    if preferred.exists():
        return preferred
    files = sorted(cwd.glob("*.xlsx"))
    if not files:
        raise FileNotFoundError("当前目录未找到 .xlsx 文件。")
    return files[0]


def choose_output_path(base_dir: Path, stock_name: str, suffix: str = ".xlsx") -> Path:
    stem = sanitize_filename(stock_name)
    out = base_dir / f"{stem}{suffix}"
    if not out.exists():
        return out
    idx = 1
    while True:
        candidate = base_dir / f"{stem}_{idx}{suffix}"
        if not candidate.exists():
            return candidate
        idx += 1


def fill_workbook(template_path: Path, values: dict[str, Any], output_path: Path) -> tuple[int, int]:
    wb = openpyxl.load_workbook(template_path)
    filled_cells = 0
    row_count = 0

    for ws in wb.worksheets:
        headers = {str(ws.cell(1, c).value).strip(): c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}
        field_col = headers.get("字段英文名", 3)
        sample_col = headers.get("示例值(留空)", 11)

        for r in range(2, ws.max_row + 1):
            field = ws.cell(r, field_col).value
            if not field:
                continue
            field_key = str(field).strip()
            value = values.get(field_key)
            if value is None or (isinstance(value, float) and (math.isnan(value) or math.isinf(value))):
                value = "N/A"
            if value == "":
                value = "N/A"
            ws.cell(r, sample_col).value = value
            filled_cells += 1
            row_count += 1

    wb.save(output_path)
    return filled_cells, row_count


def main() -> None:
    started = time.perf_counter()
    cwd = Path.cwd()
    template_path = find_template_excel(cwd)

    stock_code = input("请输入A股6位股票代码（例如 603305）: ").strip()
    collector = EastmoneyDataCollector(stock_code)
    values = collector.build_field_values()

    stock_name = str(values.get("stock_name") or stock_code)
    output_path = choose_output_path(cwd, stock_name)
    filled_cells, row_count = fill_workbook(template_path, values, output_path)

    print(f"抓取完成：{stock_name} ({stock_code})")
    print(f"模板文件：{template_path}")
    print(f"输出文件：{output_path}")
    print(f"已写入示例值单元格：{filled_cells}（字段行总数：{row_count}）")
    print(f"总耗时：{time.perf_counter() - started:.2f}s")


if __name__ == "__main__":
    main()
